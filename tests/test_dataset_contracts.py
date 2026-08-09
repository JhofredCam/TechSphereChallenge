from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.bootstrap import bootstrap as app_bootstrap
from app.dataset import (
    CLINICAL_FILE,
    CONVERSATIONS_FILE,
    DEMOGRAPHIC_FILE,
    EXPECTED_HEADERS,
    TRAJECTORIES_FILE,
    DatasetValidationError,
    iter_result_rows,
    parse_json_cell,
    validate_dataset,
    validate_dataset_strict,
    validate_joins,
    validate_workbook,
)
from scripts.validate_dataset import (
    format_validation_report,
    validation_report_to_dict,
)
from scripts.validate_dataset import (
    main as validate_dataset_main,
)


def _write_workbook(path: Path, headers: tuple[str, ...], rows: list[tuple[object, ...]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "result"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)


def test_dataset_json_cells_reject_invalid_json_and_non_lists(tmp_path):
    source = tmp_path / "clinical.xlsx"
    _write_workbook(
        source,
        ("paciente_id", "comorbilidades"),
        [("pac-1", "{not valid json}")],
    )

    report = validate_workbook(
        source,
        expected_headers=("paciente_id", "comorbilidades"),
        expected_rows=1,
        json_fields=("comorbilidades",),
    )

    assert report.valid is False  # UT-DATA-02: malformed embedded JSON is explicit.
    assert any("not valid JSON" in error for error in report.errors)
    assert parse_json_cell('["diabetes"]') == ["diabetes"]
    with pytest.raises(ValueError, match="must contain a JSON list"):
        parse_json_cell('{"condition": true}')


def test_dataset_workbook_contract_reports_sheet_header_and_count_errors(tmp_path):
    source = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "result"
    worksheet.append(["id", "id"])
    worksheet.append(["one", "duplicate"])
    workbook.create_sheet("extra")
    workbook.save(source)

    report = validate_workbook(
        source,
        expected_headers=("expected",),
        expected_rows=2,
    )

    assert report.valid is False  # UT-DATA-01: the result-sheet contract fails closed.
    assert any("exactly one sheet" in error for error in report.errors)
    assert any("headers do not match" in error for error in report.errors)
    assert any("data rows" in error for error in report.errors)


def test_dataset_iter_rows_skips_blank_rows_and_pads_short_rows(tmp_path):
    source = tmp_path / "rows.xlsx"
    _write_workbook(
        source,
        ("id", "value"),
        [("one",), (None, None), ("two", "ok")],
    )

    rows = list(iter_result_rows(source, expected_headers=("id", "value")))

    assert rows == [
        {"id": "one", "value": None},
        {"id": "two", "value": "ok"},
    ]


def test_dataset_joins_report_identity_and_layer_mismatches(tmp_path):
    errors = validate_joins(
        clinical_rows=[{"paciente_id": "pac-1"}],
        demographic_rows=[{"paciente_id": "pac-2"}],
        trajectory_rows=[
            {"trayectoria_id": "t-1", "paciente_id": "pac-2"},
            {"trayectoria_id": "t-1", "paciente_id": "pac-2"},
        ],
        conversation_rows=[
            {"caso_id": "caso_t-1", "paciente_id": "pac-1", "capa": "capa2_ruidosa"},
            {"caso_id": "caso_missing", "paciente_id": "pac-3", "capa": "unknown"},
        ],
    )

    assert "clinical and demographic paciente_id sets do not match" in errors
    assert "trajectory IDs are not unique" in errors
    assert any("unknown layer" in error for error in errors)
    assert any("patient does not match its trajectory" in error for error in errors)
    assert any("unknown patient" in error for error in errors)
    assert any("does not join a trajectory" in error for error in errors)


def test_dataset_validation_reports_missing_files_and_strict_mode_raises(tmp_path):
    report = validate_dataset(tmp_path, expected_counts=None)

    assert report.valid is False
    assert len(report.tables) == len(EXPECTED_HEADERS)
    with pytest.raises(DatasetValidationError):
        validate_dataset_strict(tmp_path)


def test_dataset_fixture_uses_the_four_named_workbooks_without_canonical_data(tmp_path):
    rows = {
        CLINICAL_FILE: {"paciente_id": "pac-1", "comorbilidades": "[]"},
        DEMOGRAPHIC_FILE: {"paciente_id": "pac-1", "adaptation_fields": "[]"},
        TRAJECTORIES_FILE: {"trayectoria_id": "t-1", "paciente_id": "pac-1"},
        CONVERSATIONS_FILE: {
            "caso_id": "caso_t-1",
            "paciente_id": "pac-1",
            "capa": "capa1_limpia",
        },
    }
    for filename, headers in EXPECTED_HEADERS.items():
        _write_workbook(
            tmp_path / filename,
            headers,
            [tuple(rows[filename].get(header) for header in headers)],
        )

    report = validate_dataset(tmp_path, expected_counts=None)

    assert report.valid is True  # IT-DATA-01 fixture: joins use IDs, never row positions.


def test_dataset_validation_cli_and_app_wrapper_use_the_same_contract(tmp_path, capsys):
    rows = {
        CLINICAL_FILE: {"paciente_id": "pac-1", "comorbilidades": "[]"},
        DEMOGRAPHIC_FILE: {"paciente_id": "pac-1", "adaptation_fields": "[]"},
        TRAJECTORIES_FILE: {"trayectoria_id": "t-1", "paciente_id": "pac-1"},
        CONVERSATIONS_FILE: {
            "caso_id": "caso_t-1",
            "paciente_id": "pac-1",
            "capa": "capa1_limpia",
        },
    }
    for filename, headers in EXPECTED_HEADERS.items():
        _write_workbook(
            tmp_path / filename,
            headers,
            [tuple(rows[filename].get(header) for header in headers)],
        )

    report = validate_dataset(tmp_path, expected_counts=None)

    assert app_bootstrap.__module__ == "scripts.bootstrap"
    assert validation_report_to_dict(report)["valid"] is True
    assert "Dataset validation: valid" in format_validation_report(report)
    # The CLI intentionally uses canonical row counts; this one-row fixture therefore
    # exercises the explicit invalid result without reading the repository dataset.
    assert validate_dataset_main(["--dataset-dir", str(tmp_path), "--json"]) == 1
    assert '"valid": false' in capsys.readouterr().out

    missing = tmp_path / "missing"
    missing.mkdir()
    assert validate_dataset_main(["--dataset-dir", str(missing)]) == 1
    assert "Dataset validation: invalid" in capsys.readouterr().out

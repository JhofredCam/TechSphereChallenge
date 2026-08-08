"""Validation helpers for the four local challenge workbooks.

The XLSX files are intentionally read in read-only mode and row-by-row.  The validator
does not reconstruct conversations or write a normalized copy, which avoids mixing the
two conversation layers and keeps the canonical dataset untouched.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

from .schemas import DatasetTableReport, DatasetValidationReport

RESULT_SHEET = "result"

CONVERSATIONS_FILE = "dataset_final.xlsx"
CLINICAL_FILE = "perfiles_clinicos_pacientes_silver_contest.xlsx"
DEMOGRAPHIC_FILE = "perfiles_pacientes_co.xlsx"
TRAJECTORIES_FILE = "trayectorias_postop_silver.xlsx"

EXPECTED_ROW_COUNTS = {
    CONVERSATIONS_FILE: 3991,
    CLINICAL_FILE: 40,
    DEMOGRAPHIC_FILE: 40,
    TRAJECTORIES_FILE: 160,
}

EXPECTED_HEADERS: dict[str, tuple[str, ...]] = {
    CONVERSATIONS_FILE: (
        "dialogo_id",
        "caso_id",
        "paciente_id",
        "dia_postop",
        "turno_idx",
        "hablante",
        "texto",
        "label_ground_truth",
        "estilo_paciente",
        "modelo_paciente",
        "modelo_agente",
        "capa",
        "generado_ts",
    ),
    CLINICAL_FILE: (
        "paciente_id",
        "bundle_id",
        "synthea_runtime",
        "modulo_synthea",
        "procedimiento",
        "fecha_cirugia",
        "edad",
        "genero",
        "comorbilidades",
        "complicacion_encounter",
        "generado_ts",
    ),
    DEMOGRAPHIC_FILE: (
        "paciente_id",
        "nombre_completo",
        "direccion",
        "ciudad",
        "departamento",
        "documento_cc",
        "eps",
        "source_country",
        "adapted_country",
        "adaptation_fields",
        "adaptation_ts",
    ),
    TRAJECTORIES_FILE: (
        "trayectoria_id",
        "paciente_id",
        "dia_postop",
        "arquetipo_trayectoria",
        "dolor_nrs",
        "fiebre_c",
        "movilidad",
        "herida",
        "apetito",
        "sueno",
        "seed",
        "generado_ts",
    ),
}

JSON_FIELDS: dict[str, tuple[str, ...]] = {
    CLINICAL_FILE: ("comorbilidades",),
    DEMOGRAPHIC_FILE: ("adaptation_fields",),
}


class DatasetDependencyError(RuntimeError):
    pass


class DatasetValidationError(ValueError):
    def __init__(self, report: DatasetValidationReport) -> None:
        self.report = report
        message = "; ".join(report.errors) or "dataset validation failed"
        super().__init__(message)


@dataclass(frozen=True, slots=True)
class _LoadedTable:
    report: DatasetTableReport
    rows: tuple[dict[str, Any], ...]


def _open_workbook(path: Path) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DatasetDependencyError(
            "openpyxl is required to validate the challenge workbooks"
        ) from exc
    return load_workbook(path, read_only=True, data_only=True)


def _header_names(values: Sequence[Any]) -> tuple[str, ...]:
    return tuple("" if value is None else str(value).strip() for value in values)


def _validate_json_value(value: Any, field_name: str) -> Any:
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{field_name} is not valid JSON: {exc.msg}") from exc
    else:
        parsed = value
    if not isinstance(parsed, list):
        raise ValueError(f"{field_name} must contain a JSON list")
    return parsed


def parse_json_cell(value: Any, *, field_name: str = "JSON field") -> list[Any]:
    """Parse and validate one embedded JSON-list cell."""

    return _validate_json_value(value, field_name)


def _iter_result_rows_with_numbers(
    path: str | Path,
    *,
    expected_headers: Sequence[str] | None = None,
) -> Iterator[tuple[int, dict[str, Any]]]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(source)
    workbook = _open_workbook(source)
    try:
        if workbook.sheetnames != [RESULT_SHEET]:
            raise ValueError(
                f"{source.name} must contain only the {RESULT_SHEET!r} sheet; "
                f"found {workbook.sheetnames!r}"
            )
        worksheet = workbook[RESULT_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = _header_names(next(rows))
        except StopIteration as exc:
            raise ValueError(f"{source.name} has no header row") from exc
        if not headers or any(not header for header in headers):
            raise ValueError(f"{source.name} has an empty header")
        if len(set(headers)) != len(headers):
            raise ValueError(f"{source.name} has duplicate headers")
        if expected_headers is not None and tuple(expected_headers) != headers:
            raise ValueError(
                f"{source.name} headers do not match the expected result contract"
            )

        for row_number, values in enumerate(rows, start=2):
            cells = tuple(values)
            if not cells or all(value is None for value in cells):
                continue
            padded = cells + (None,) * max(0, len(headers) - len(cells))
            yield row_number, dict(zip(headers, padded[: len(headers)]))
    finally:
        workbook.close()


def iter_result_rows(
    path: str | Path,
    *,
    expected_headers: Sequence[str] | None = None,
) -> Iterator[dict[str, Any]]:
    """Yield non-empty rows from the sole ``result`` sheet as dictionaries."""

    for _, row in _iter_result_rows_with_numbers(path, expected_headers=expected_headers):
        yield row


def _load_table(
    path: Path,
    *,
    expected_headers: Sequence[str] | None = None,
    expected_rows: int | None = None,
    json_fields: Iterable[str] = (),
) -> _LoadedTable:
    errors: list[str] = []
    rows: list[dict[str, Any]] = []
    headers: tuple[str, ...] = ()
    sheet_name: str | None = None
    if not path.is_file():
        return _LoadedTable(
            DatasetTableReport(
                path=str(path),
                sheet_name=None,
                headers=(),
                row_count=0,
                errors=(f"missing workbook: {path}",),
            ),
            (),
        )

    try:
        workbook = _open_workbook(path)
    except Exception as exc:
        return _LoadedTable(
            DatasetTableReport(
                path=str(path),
                sheet_name=None,
                headers=(),
                row_count=0,
                errors=(str(exc),),
            ),
            (),
        )
    try:
        if RESULT_SHEET not in workbook.sheetnames:
            errors.append(f"{path.name} is missing the {RESULT_SHEET!r} sheet")
        if workbook.sheetnames != [RESULT_SHEET]:
            errors.append(
                f"{path.name} must contain exactly one sheet named {RESULT_SHEET!r}"
            )
        if RESULT_SHEET in workbook.sheetnames:
            sheet_name = RESULT_SHEET
            worksheet = workbook[RESULT_SHEET]
            row_iterator = worksheet.iter_rows(values_only=True)
            try:
                headers = _header_names(next(row_iterator))
            except StopIteration:
                errors.append(f"{path.name} has no header row")
                row_iterator = iter(())
            if headers:
                if any(not header for header in headers):
                    errors.append(f"{path.name} has an empty header")
                if len(set(headers)) != len(headers):
                    errors.append(f"{path.name} has duplicate headers")
                if expected_headers is not None and tuple(expected_headers) != headers:
                    errors.append(f"{path.name} headers do not match the expected result contract")
                json_field_set = tuple(json_fields)
                for row_number, values in enumerate(row_iterator, start=2):
                    cells = tuple(values)
                    if not cells or all(value is None for value in cells):
                        continue
                    padded = cells + (None,) * max(0, len(headers) - len(cells))
                    row = dict(zip(headers, padded[: len(headers)]))
                    rows.append(row)
                    for field_name in json_field_set:
                        if field_name not in row:
                            errors.append(f"{path.name} has no {field_name!r} column")
                            continue
                        try:
                            _validate_json_value(row[field_name], field_name)
                        except ValueError as exc:
                            errors.append(f"{path.name} row {row_number}: {exc}")
    finally:
        workbook.close()

    if expected_rows is not None and len(rows) != expected_rows:
        errors.append(
            f"{path.name} has {len(rows)} data rows; expected {expected_rows}"
        )
    report = DatasetTableReport(
        path=str(path),
        sheet_name=sheet_name,
        headers=headers,
        row_count=len(rows),
        errors=tuple(errors),
    )
    return _LoadedTable(report, tuple(rows))


def validate_workbook(
    path: str | Path,
    *,
    expected_headers: Sequence[str] | None = None,
    expected_rows: int | None = None,
    json_fields: Iterable[str] = (),
) -> DatasetTableReport:
    """Validate one workbook's sheet, headers, row count, and JSON cells."""

    return _load_table(
        Path(path),
        expected_headers=expected_headers,
        expected_rows=expected_rows,
        json_fields=json_fields,
    ).report


def _duplicate_values(rows: Iterable[Mapping[str, Any]], field: str) -> set[Any]:
    seen: set[Any] = set()
    duplicates: set[Any] = set()
    for row in rows:
        value = row.get(field)
        if value is None:
            continue
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return duplicates


def validate_joins(
    clinical_rows: Iterable[Mapping[str, Any]],
    demographic_rows: Iterable[Mapping[str, Any]],
    trajectory_rows: Iterable[Mapping[str, Any]],
    conversation_rows: Iterable[Mapping[str, Any]],
) -> tuple[str, ...]:
    """Validate patient and case joins without combining conversation layers."""

    clinical = tuple(clinical_rows)
    demographic = tuple(demographic_rows)
    trajectories = tuple(trajectory_rows)
    conversations = tuple(conversation_rows)
    errors: list[str] = []

    clinical_ids = {row.get("paciente_id") for row in clinical}
    demographic_ids = {row.get("paciente_id") for row in demographic}
    clinical_ids.discard(None)
    demographic_ids.discard(None)
    if clinical_ids != demographic_ids:
        errors.append("clinical and demographic paciente_id sets do not match")

    duplicate_trajectory_ids = _duplicate_values(trajectories, "trayectoria_id")
    if duplicate_trajectory_ids:
        errors.append("trajectory IDs are not unique")
    trajectory_by_id: dict[Any, Mapping[str, Any]] = {}
    for row in trajectories:
        trajectory_id = row.get("trayectoria_id")
        patient_id = row.get("paciente_id")
        if patient_id not in clinical_ids:
            errors.append(f"trajectory {trajectory_id!r} references an unknown patient")
        if trajectory_id is not None:
            trajectory_by_id[trajectory_id] = row

    allowed_layers = {"capa1_limpia", "capa2_ruidosa"}
    for row in conversations:
        case_id = row.get("caso_id")
        patient_id = row.get("paciente_id")
        layer = row.get("capa")
        if layer not in allowed_layers:
            errors.append(f"conversation {case_id!r} has an unknown layer {layer!r}")
        if patient_id not in clinical_ids:
            errors.append(f"conversation {case_id!r} references an unknown patient")
        expected_case = None
        if isinstance(case_id, str) and case_id.startswith("caso_"):
            expected_case = case_id.removeprefix("caso_")
        trajectory = trajectory_by_id.get(expected_case)
        if trajectory is None:
            errors.append(
                f"conversation {case_id!r} does not join a trajectory as caso_<trayectoria_id>"
            )
        elif trajectory.get("paciente_id") != patient_id:
            errors.append(f"conversation {case_id!r} patient does not match its trajectory")

    return tuple(dict.fromkeys(errors))


def validate_dataset(
    dataset_dir: str | Path,
    *,
    expected_counts: Mapping[str, int | None] | None = EXPECTED_ROW_COUNTS,
    raise_on_error: bool = False,
) -> DatasetValidationReport:
    """Validate all four workbooks and their cross-file joins."""

    root = Path(dataset_dir)
    tables: list[DatasetTableReport] = []
    loaded: dict[str, _LoadedTable] = {}
    for filename, headers in EXPECTED_HEADERS.items():
        expected_rows = None if expected_counts is None else expected_counts.get(filename)
        table = _load_table(
            root / filename,
            expected_headers=headers,
            expected_rows=expected_rows,
            json_fields=JSON_FIELDS.get(filename, ()),
        )
        loaded[filename] = table
        tables.append(table.report)

    errors = [error for table in tables for error in table.errors]
    required = (CLINICAL_FILE, DEMOGRAPHIC_FILE, TRAJECTORIES_FILE, CONVERSATIONS_FILE)
    if all(loaded[name].rows for name in required):
        errors.extend(
            validate_joins(
                loaded[CLINICAL_FILE].rows,
                loaded[DEMOGRAPHIC_FILE].rows,
                loaded[TRAJECTORIES_FILE].rows,
                loaded[CONVERSATIONS_FILE].rows,
            )
        )
    report = DatasetValidationReport(
        tables=tuple(tables),
        errors=tuple(dict.fromkeys(errors)),
    )
    if raise_on_error and not report.valid:
        raise DatasetValidationError(report)
    return report


def validate_dataset_strict(dataset_dir: str | Path) -> DatasetValidationReport:
    return validate_dataset(dataset_dir, raise_on_error=True)


# Names used by bootstrap and focused tests can remain concise without duplicating logic.
validate_xlsx = validate_workbook
iter_xlsx_rows = iter_result_rows
validate_all = validate_dataset


__all__ = [
    "CLINICAL_FILE",
    "CONVERSATIONS_FILE",
    "DEMOGRAPHIC_FILE",
    "EXPECTED_HEADERS",
    "EXPECTED_ROW_COUNTS",
    "JSON_FIELDS",
    "RESULT_SHEET",
    "TRAJECTORIES_FILE",
    "DatasetDependencyError",
    "DatasetValidationError",
    "iter_result_rows",
    "iter_xlsx_rows",
    "parse_json_cell",
    "validate_all",
    "validate_dataset",
    "validate_dataset_strict",
    "validate_joins",
    "validate_workbook",
    "validate_xlsx",
]
